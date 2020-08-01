#====tkinter框架、非界面框架
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import *
from tkinter import *
from openpyxl import load_workbook
import tkinter.messagebox
import time
import os
import pandas as pd
import numpy as np
import sqlite3
import openpyxl
import hashlib 
import logging 
import sys 
import re
import math
import xlrd,xlwt
import codecs
import chardet
import winreg
import win32api,win32gui,win32con,win32com,win32print
import win32com.client as win32	
import filecmp




#*********************************************************************
def 无索引取数(list='',datapath="",index_name=None,dest='e:\\desktop\\提取.xlsx'): #list是所查的内容；index_name是列名; datapath数据文件夹; dest 目标文件名
	if list=='': 
		file=tkinter.filedialog.askopenfilename() #索引文件
		list=excel_col_to_list(file)
	elif os.path.exists(list)==True: list=excel_col_to_list(list) #list是列表就直接使用，是包含列表的文件就读入它；
	
	files=list_all_files(datapath)
	#日志(list)
	yy=1 #destBook当前行数控制  
	oExcel=vbaInit()
	destBook=oExcel.Workbooks.Add()
	
	for ffile in files:
		try: 
			srcBook=oExcel.Workbooks.Open(ffile)
		except:
			continue
		for sht in srcBook.Sheets:
			s_col=sht.UsedRange.Columns.Count
			s_row=0;
			for lis in list:
				ss=sht.Cells.Find(lis)
				while ss!=None and ss.Row>s_row:
					s_row=ss.Row
					ss=sht.Cells.FindNext(ss)
					sht.Rows(s_row).Copy(destBook.ActiveSheet.Cells(yy,1))
					destBook.ActiveSheet.Cells(yy,s_col+1).Value=lis
					destBook.ActiveSheet.Cells(yy,s_col+2).Value=ffile #写入文件路径备用
					yy+=1
					#if yy==1000000: destBook.SaveAs(dest)  #加入100万行自动保存的语句??
					#日志(str(ss.Row))
		srcBook.Close()	
	destBook.SaveAs(dest)
#*******************************************************************
def 查询文件夹_to_csv(list='',datapath="",index_name=None,dest='提取.csv'): #list是所查的内容；index_name是列名; datapath数据文件夹; dest 目标文件名
	if list=='': 
		file=tkinter.filedialog.askopenfilename() #索引文件
		list=excel_col_to_list(file)
	elif os.path.exists(list)==True: list=excel_col_to_list(list) #list是列表就直接使用，是包含列表的文件就读入它；
	
	files=list_all_files(datapath)
	#日志(files)
	
	ansi_to_utf8(datapath)
	hh=1
	for file in files:
		if file.endswith(".xls") or file.endswith(".xlsx"): df=pd.read_excel(file)  
		elif file.endswith(".csv"):  df=pd.read_csv(file)
		if index_name!=None:  set_index=df.set_index(index_name)  #做索引提升速度；None不行吧??
		else: set_index=df
		
		#print(df.columns.values) #列名
		for lis in list:
			if lis in set_index.index:
				ggg=set_index.loc[lis]
				if isinstance(ggg, pd.Series):
					ggg=set_index.loc[[lis]]  #Series转置为DataFrame格式
				ggg.to_csv(dest,mode='a',encoding="ansi",header=hh)
				hh=0 if hh==1 else 0

#*******************************************************************
def 获取文件编码(file=''):
	if file=='': file=askopenfilename()
	if os.path.isfile(path):
		f = open(file,'rb')
		data = f.read()
		return chardet.detect(data).get("encoding") 
		#return chardet.detect(data) #给出完整信息
		#return chardet.detect(data)['encoding']

#********************************************************************
#转换为utf8
def ansi_to_utf8(file_path=""):
	if file_path=="":  file_path =askdirectory()
	files = list_all_files(file_path)
	#日志(files)

	for file in files:
		if file.endswith(".xls") or file.endswith(".xlsx"):  continue
		if 获取文件编码(file)=='GB2312':
			f = codecs.open(file, 'r', 'ansi')
			ff = f.read()
			file_object = codecs.open(file, 'w', 'utf-8')
			file_object.write(ff)


#********************************************************************
#转换单个utf8文件为ANSI
def utf8_to_ansi(utf8file,ansifile):
	if 获取文件编码(utf8file)=='utf-8':
		f = codecs.open(utf8file,'r','utf8')
		utfstr = f.read()
		f.close()
		#把UTF8字符串转码成ANSI字符串
		ansiStr = utfstr.encode('mbcs')
		#使用二进制格式保存转码后的文本
		f = open(ansifile,'wb')
		f.write(ansiStr)
		f.close()

#*******************************************************************
def 批量utf8_to_ansi(path=""):
	files=list_all_files(path) #???换函数
	for file in files:
		if 获取文件编码(file)=='utf-8':
			utf8_To_Ansi(file,file)




#*********************************************************************
#透视表
def 透视表(table=""):
	if table=="":
		table=openDialog()
	df = pd.read_excel(table)
	res = pd.pivot_table(df,index=[u'交易对手名称'],values=[u'发生额'],aggfunc=[np.sum]).to_excel("透视_.xlsx")
	print(res)
	df.head()

#**********************************************************************
def 一键出报告():
	解压目录(folder)
	删除重复文件(path=folder)
	批量转ansi(path="")
	合并(folder)
	

#初始化
def vbaInit(): 
  if not 'oExcel' in dir():
    oExcel=win32com.client.Dispatch("Excel.Application")
    oExcel.Visible=True
    oExcel.Application.DisplayAlerts = False
  return oExcel


#通用于打开txt、excel文件；返回打开的文件对象
def openEx(filePath=""):
	fso=win32com.client.Dispatch("Scripting.FileSystemObject")
	#tkinter.messagebox.showinfo(filePath)
	oExcel=vbaInit()
	if filePath=="":
		filePath=oExcel.Application.FileDialog(1)
		
		filePath.InitialFileName = "E:\desktop" #??动态获取桌面路径
		filePath.Filters.Add("什么情况","*.xls*; *.txt",1) 
		filePath.Show()
		filePath=filePath.SelectedItems(1) # //对象最后又改成了纯路径
	if fso.GetExtensionName(filePath)=="txt":
		dest=fso.OpenTextFile(filePath,1)	#stream=fso.GetFile(filePath).OpenAsTextStream()
	else:
		dest=oExcel.Workbooks.Open(filePath)
	return dest


#获取表头所在的行; 根据参数2的数组内容模糊匹配获取表头行；参数2缺省时也应能判断，因为有时没有表头。后续开发???
def 取表头行(sht="",biaotou=""):
	bv=0
	maxv=0
	srow=0
	for row in range(1,15):#row<15,控制在15行之内进行探测，一般是够用的
		for bi in range(0,biaotou.length):
			sss=xlsFindReg(biaotou[bi],sht,row)
			if not sss==null:
				bv=bv+1 #bi控制表头数量
		#for(bi=0;bi<5;bi++){sss=xlsFindReg(biaotou[bi],sht,row);if(sss!=null)bv++} #bi控制表头数量
		if bv>maxv:
			maxv=bv
			srow=row #取匹配数最多的这一行做为表头的所在行
		bv=0  #获取取表头行
		return srow
#***************************************************************************************
#读取xls
def read_excel_xls(path):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    for i in range(0, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
        print()

#***************************************************************************************
#覆盖写入xls
def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")

#***************************************************************************************
#覆盖写入xlsx
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")
    
#***************************************************************************************
#添加写入xls
def write_excel_xls_append(path, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            new_worksheet.write(i+rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
    new_workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")

#***************************************************************************************
#添加写入xlsx
def write_excel_xlsx_append(path, value):
	import openpyxl
	data = openpyxl.load_workbook('excel_test.xlsx')
	print(data.get_named_ranges()) # 输出工作页索引范围
	print(data.get_sheet_names()) # 输出所有工作页的名称
	# 取第一张表
	sheetnames = data.get_sheet_names()
	table = data.get_sheet_by_name(sheetnames[0])
	table = data.active
	print(table.title) # 输出表名
	nrows = table.max_row # 获得行数
	ncolumns = table.max_column # 获得行数
	values = ['E','X','C','E','L']
	for value in values:
	    table.cell(nrows+1,1).value = value
	    nrows = nrows + 1
	data.save('excel_test.xlsx')

 
#***************************************************************************************
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

#****批量打印********************************************************************
def 批量打印(file_dirs=""):
	file_dirs = list_all_files(file_dirs) 
	b=len(file_dirs)
	i = 0
	while i < len(file_dirs):
		ext = os.path.splitext(file_dirs[i])[1]
		if ext.startswith('.x'):
            # excel
			xlApp = win32com.client.Dispatch('Excel.Application')
			xlApp.Visible = 0
			xlApp.EnableEvents = False
			xlApp.DisplayAlerts = False
			xlBook = xlApp.Workbooks.Open(file_dirs[i])
			xlApp.ActiveWorkbook.Sheets(1).PageSetup.Zoom = False
			xlApp.ActiveWorkbook.Sheets(1).PageSetup.FitToPagesWide = 1
			xlApp.ActiveWorkbook.Sheets(1).PageSetup.FitToPagesTall = 1
			xlBook.PrintOut(1, 99, )
			xlApp.quit()
		else:
            # word pdf txt
			win32api.ShellExecute(0,"print",file_dirs[i],'/d:"%s"' % win32print.GetDefaultPrinter(),".",0)
		print(file_dirs[i])
		time.sleep(1)
		i = i + 1


#**************************************************************************************
#正则式查找，找到返回单元格对象; 最小查找是行或列，单元格不用函数可直接用正则式操作；
#ranges:默认全表，>=1行，2就是第二行，<=-1列，-3就是第三列；xfan=1时，从右往左找; yfan=1时，从后往前查找; 
def xlsFindReg(reg,sht,ranges,xfan,yfan):
	#var i,sss,c={},y1,yEnd,yp,x1,xEnd,xp
	col=sht.UsedRange.Columns.Count
	row=sht.UsedRange.rows.count
	if yfan==1:
		y1=row
		yEnd=1
		yp=-1
	else:
		y1=1
		yEnd=row
		yp=1
	if xfan==1:
		x1=col
		xEnd=1
		xp=-1
	else:
		x1=1
		xEnd=col
		xp=1
	if ranges>=1:
		y1=ranges
		yEnd=ranges+1
		yp=1
	elif ranges<0:
		x1=-ranges
		xEnd=-ranges+1
		xp=1
	for c.y in ranges(y1,yEnd,yp):
		for c.x in ranges(x1,xEnd,xp):
			sss=String(sht.cells(c.y,c.x).value).replace(r'/\s/g',"")
		if reg.test(sss):
			return sht.cells(c.y,c.x)	
	return null


#*****************************************************************************
#编辑列，令其整齐一致，但是尽量不删任何列；该表用于提取数据
def 表次序整理(dataBookpath="",sht=""):
	oExcel=vbaInit()
	Ws=win32com.client.Dispatch("WScript.shell")
	if dataBookpath=="":
		dataBookpath=Workpath+"流水\\"
	if not os.path.exists(dataBookpath+"整理\\"):
		os.makedirs(dataBookpath+"整理\\")
	fnames=list_all_files(dataBookpath)
	
	for loop in fnames: #这是一个文件夹的循环
		data=myOpen(loop) #银行流水
		if(fso().GetExtensionName(data.name)=="xls"):
			lns=65535
		else:
			lns=1048576 #数据txt\html咋办???
		if sht=="":
			shtBegin=1
			shtEnd=data.Sheets.count
		else:
			shtBegin=sht
			shtEnd=sht
		for si in range(shtBegin,shtEnd): #这是一个工作簿的循环
			mysht=data.Sheets(si)
			cells=mysht.cells
			mysht.Select()
			Ws.Popup("整理第 "+si+" 个表","1","阿弥佗佛","0") 
			srow=取表头行(mysht,r'[/日期/,/[姓户]名/,/[卡帐账][号户]/,/对[方手]/,/[金余]额|发生额/]')
			if xlsFindReg(r'日期',mysht,srow)>0:
				cut(mysht,xlsFindReg(r'日期',mysht,srow).column,1)
			elif xlsFindReg(r'时间',mysht,srow)>0: #??>0对不对？？
				cut(mysht,xlsFindReg(r'时间',mysht,srow).column,1)
			else:
				mysht.columns(1).insert() #纠错之用：找不到就查个空列占位
			if xlsFindReg(r'^(?!对方).*([姓户]名|客户名?$)',mysht,srow):
				cut(mysht,xlsFindReg(r'^(?!对方).*([姓户]名|客户名?$)',mysht,srow).column,2)
			else:
				mysht.columns(2).insert() #纠错之用：找不到就查个空列占位
			if xlsFindReg(r'^(?!对方).*[卡帐账][号户]',mysht,srow):
				cut(mysht,xlsFindReg(r'^(?!对方).*[卡帐账][号户]',mysht,srow).column,3)
			else:
				mysht.columns(3).insert()
			if xlsFindReg(r'对[方手].*[账帐卡][户号]$',mysht,srow):
				cut(mysht,xlsFindReg(r'对[方手].*[账帐卡][户号]$',mysht,srow).column,4)
			elif xlsFindReg(r'[账帐卡][户号]$',mysht,srow,1):
				cut(mysht,xlsFindReg(r'[账帐卡][户号]$',mysht,srow,1).column,4)
			else:
				mysht.columns(4).insert()
			if xlsFindReg(r'对[方手].*名',mysht,srow):
				cut(mysht,xlsFindReg(r'对[方手].*名',mysht,srow).column,5)
			elif xlsFindReg(r'户名|名称|姓名',mysht,srow,1):
				cut(mysht,xlsFindReg(r'户名|名称|姓名',mysht,srow,1).column,5)
			else:
				mysht.columns(5).insert()
			if xlsFindReg(r'借方|付款',mysht,srow) and xlsFindReg(r'贷方|收款',mysht,srow): #对付金额分成借贷两列的情况
				mysht.columns(6).insert()
				sss=xlsFindReg(r'借方|付款',mysht,srow)
				x=sss.column
				cells(srow,6).value="交易金额"  #写表头字段
				for i in range(srow+1,mysht.Range("A"+lns).End(-4162).Row):
					if cells(i,x).value>0:
						cells(i,6).value=-cells(i,x).value
				sss=xlsFindReg(r'贷方|收款',mysht,srow)
				x=sss.column
				for i in range(srow+1,mysht.Range("A"+lns).End(-4162).Row):
					if cells(i,x).value!=0:
						cells(i,6).value=--cells(i,x) #--是为了解决文本类型数字的问题
			elif xlsFindReg(r'金额|发生额',mysht,srow): #此是仅有一列金额的情况，以"金额|发生额"定位此列
				cut(mysht,xlsFindReg(r'金额|发生额',mysht,srow.column,6))  #此列剪切至第6列
				col=xlsFindReg(r'借贷标志|收付标志|进出|交易方向|交易类型',mysht,srow)
				
				for i in range(srow+1,mysht.Range("A"+lns).End(-4162).Row):
					if r'借|出|付|D|取现|现取|提现|消费|0011|0001|^[0]$|退'.match(cells(i,col.column).value)!=0 and cells(i,6).value>0:
						cells(i,6).value=-math.fabs(cells(i,6).value)  #单列的数字有时有正负号，有时候没有。故此不加判断直接给与绝对负数
					else:
						cells(i,6).value=--cells(i,6).value #负负得正转换文本数字				
			else:
				mysht.columns(6).insert() #找不到这列，就插入空列
			if cells(srow+1,3).value==null:
				mysht.Delete()
				shtEnd=shtEnd-1
				si=si-1  #删除空表
			else:
				mysht.name=String(cells(srow+1,2).value).substr(0,3)+String(cells(srow+1,3).value).slice(-4)#??表名应该动态，3是目前的选定

			##########一个工作簿的循环尾
		fpath=Workpath+"流水\\整理\\整理"+loop
		data.close(true,fpath)
		arr.push(fpath)
		#大循环
	return arr

#*****************************************************************************
def 一键打印():
    os.system('WScript mylib\\Excel一键打印.js')


#*****************************************************************************
#合并为一个sheet 
def 合并为csv(folder="",destpath="合并.csv"):
	pd.DataFrame().to_csv(destpath)
	df1=pd.read_csv(destpath)
	if folder=="":
		folder=askdirectory()
	files=list_all_files(folder)
	tkinter.messagebox.showinfo("",files)
	for i in files:
		df2=pd.read_excel(i)
		#df1=df1.append(df2)
		df2.to_csv(destpath,mode='a',encoding="utf_8")
	tkinter.messagebox.showinfo("","ok")

#*******************************************************************
def xls_to_xlsx(rootdir=""):
    if rootdir=="":  rootdir=askdirectory()
    files = os.listdir(rootdir)
    num = len(files)
    for i in range(num):
        path = rootdir + files[i]
        a = files[i].split('.')
        b = rootdir + a[0] + '.xlsx'
        x = pd.read_excel(path)
        x.to_excel(b, index=False)


#****把txt文件转excel文件; 适用于非csv格式的txt，需要优化分隔符的设定??
def txt_xls(filename, xlsname):
    try:
        f = open(filename, 'r', encoding='utf-8')
        xls = xlwt.Workbook()
        sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
        x = 0
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break
            for i in range(len(line.split('\t'))):
                item = line.split('\t')[i]
                sheet.write(x, i, item)
            x += 1
        f.close()
        xls.save(xlsname)  # 保存xls文件
    except:
        raise

#******************************************************************************
#转换txt,csv,xls为xlsx;
def to_xlsx(folder=""):  #folder须为绝对路径
	#excel = win32.gencache.EnsureDispatch('Excel.Application') 
	excel=win32com.client.Dispatch("Excel.Application")
	excel.Visible=True
	excel.Application.DisplayAlerts = False
	fso=win32com.client.Dispatch("Scripting.FileSystemObject")
	
	if folder=="":  folder=askdirectory()
	files=list_all_files(folder)
	
	for fname in files:	
		if fname.endswith(".csv"):  utf8_to_ansi(fname,fname)
		wb = excel.Workbooks.Open(fname)
		fname=os.path.splitext(fname)[0]  #fname=os.path.basename(fname)
		fname=fname.replace('/','\\')
		#tkinter.messagebox.showinfo("",fname)
		wb.SaveAs(fname, FileFormat = 51)	#FileFormat = 51 is for .xlsx extension #FileFormat = 56 is for .xls extension
		wb.Close()
	excel.Application.Quit()

#*******************************************************************************
#合并工作簿（pd版） 
def 合并工作簿(folder="",dest="合并.xlsx"):
	writer=pd.ExcelWriter(dest)
	if folder=="":
		folder=askdirectory()
	files=list_all_files(folder)
	for file in files:
		sht=pd.ExcelFile(file) #读入原表的最快方法
		for sht_name in sht.sheet_names:
			table=sht.parse(sht_name)
			table.to_excel(writer,sheet_name=sht_name,index=False)
	writer.save()

#合并到一个工作簿 VBA版
def 合并工作簿2(folder="",dest="e:\desktop\合并.xlsx"):
	oExcel=win32com.client.Dispatch("Excel.Application")
	oExcel.Visible=True
	oExcel.Application.DisplayAlerts = False
	
	destBook=oExcel.WorkBooks.Add()
	if folder=="":
		folder=askdirectory()
	pathArr=list_all_files(folder)
	for book in pathArr:
		srcBook=oExcel.WorkBooks.Open(book)
		for i in srcBook.Sheets:
			i.Copy(destBook.Sheets(1))
		srcBook.close  #不能加括号，加上就出错
	destBook.SaveAs(dest)
	destBook.close
	oExcel.quit()
	return dest





def 索引取数():#索引数据在excel文件的 A列；
    df=pd.read_excel(tkinter.filedialog.askopenfilename(initialfile="点选索引文件"),header = None)
    dbfile = tkinter.filedialog.askopenfilename(initialfile="点选数据文件")
    conn = sqlite3.connect(dbfile)
    cur = conn.cursor()
    #cur.execute("create index if not exists index_name on user('name')")
    cur.execute('create table if not exists 成果 as select * from user where 1=0')
    for i in df.index:#根据名单进行循环
        cur.execute("insert into 成果 select * from user where name ='"+df.iat[i,0]+"'")
    cur.close()
    conn.commit()
    conn.close()
    tkinter.messagebox.showinfo("ok!")


def readDB(): #打开数据表，默认打开第一张表
    dbfile = tkinter.filedialog.askopenfilename(initialfile="点选数据文件")
    conn = sqlite3.connect(dbfile)
    #sql = "select * from user"
    sql="select name from sqlite_master"
    df = pd.read_sql(sql, conn)
    print(df)
    xz=input("请键入数字选项...")
    biao=df.iat[int(xz),0]
    sql = "select * from "+biao
    df = pd.read_sql(sql, conn)
    t1=tkinter.Text(wintk,width = 10,height = 1).grid(row=0,column=1)
    print(df)
    

##转换txt文档为数据库文件
def txt2DB():
    a=time.time()
    txtfile = tkinter.filedialog.askopenfilename(initialfile="此处数据文件名")
    file = open(txtfile, 'r',encoding='UTF-8')
    text_line = file.readline()
    columns=text_line.split('|')
    file=pd.read_csv(txtfile,sep='|',names=columns)
    txtBasename=os.path.basename(txtfile).split('.')[0] 
    file.to_sql(con = sqlite3.connect(txtBasename+'.db'), name = txtBasename,if_exists='replace',index = False)
    #tkinter.messagebox.showinfo("","耗时："+str(time.time()-a))
    return txtBasename+'.db'

    
#====tkinter框架================================================================
#****窗口类3********************************************************************
class tk_UI(tk.Frame):
	
	def __init__(self, master=None):  
		tk.Frame.__init__(self, master)
		
		screenW=master.winfo_screenwidth()#屏宽，win32api.GetSystemMetrics (0)可不依赖tk
		screenH=master.winfo_screenheight()#屏高
		winW=master.winfo_screenwidth()/2 #窗宽取屏宽的1/2
		winH=master.winfo_screenheight()*0.6 #窗高
		master.geometry("%dx%d+%d+%d"%(winW, winH, (screenW - winW)/2, (screenH - winH)/2))  #窗口在屏幕中央
		master.update()
		
		'''
		self.create_menu(root)
		self.create_content(root)
		self.path = 'C:'
		root.title("磁盘文件搜索工具")
		root.update()
		# root.resizable(False, False) 调用方法会禁止根窗体改变大小
		#以下方法用来计算并设置窗体显示时，在屏幕中心居中
		curWidth = root.winfo_width()  # get current width
		curHeight = root.winfo_height()  # get current height
		scnWidth, scnHeight = root.maxsize()  # get screen width and height
		tmpcnf = '+%d+%d' % ((scnWidth - curWidth) / 2, (scnHeight - curHeight) / 2)
		root.geometry(tmpcnf)
		root.mainloop()
		'''
	def statusBar(self,root,str):
		status = Label(root,anchor=W)  #anchor left align W -- WEST
		status.pack(side=BOTTOM,fill=X)
		sta1 = Label(status,text=str,anchor=W,width=50)  #anchor left align W -- WEST
		sta1.grid(row=0,column=0) 
		sta2 = Label(status,text='22222',anchor=W,width=500)  #anchor left align W -- WEST
		sta2.grid(row=0,column=1) 
#*************************************************************************
def statusBar(titlename,str):
	root = win32gui.FindWindow(0, titlename)
	
	status = Label(root,anchor=W)  #anchor left align W -- WEST
	status.pack(side=BOTTOM,fill=X)
	sta1 = Label(status,text=str,relief=RAISED,anchor=W,width=50)  #anchor left align W -- WEST
	sta1.grid(row=0,column=0) 
	sta2 = Label(status,text='22222',anchor=W)  #anchor left align W -- WEST
	sta2.grid(row=0,column=1) 		
#*******************************************************************************
#创建菜单
def myMenu(menudf2,win): #menuDict为列表内嵌套字典
    menubar=Menu(win) #win使用了全局变量，似应改变之?? 是否应写出通用的界面代码，自动匹配多种GUI框架
    win['menu']=menubar
    for Dict in menuDict:
        gmenu=Menu(menubar)
        for key in Dict:
            menubar.add_cascade(label=key,menu=gmenu) #第一个key添加为子菜单
            break
        for key,values in Dict.items():
            gmenu.add_command(label=key,command=values) #子菜单添加菜单项
#*****************************************************************************
def myMenu2(parent,name,dicts):
  gmenu=Menu(parent)
  parent.add_cascade(label=name,menu=gmenu)
  for key,values in dicts.items():
    gmenu.add_command(label=key,command=values) #子菜单添加菜单项
  return gmenu
	
#****对话框********************************************************************
def tkecho(headline='ok',str1='ok'):
    str1=time.localtime(time.time()).tm_year+10153917
    tkinter.messagebox.showinfo(headline,str1)

#****窗口类***************************************************************************
class AppUI():
    def __init__(self):
        root = Tk()
        self.create_menu(root)
        self.create_content(root)
        self.path = 'C:'
        root.title("磁盘文件搜索工具")
        root.update()
        # root.resizable(False, False) 调用方法会禁止根窗体改变大小
        #以下方法用来计算并设置窗体显示时，在屏幕中心居中
        curWidth = root.winfo_width()  # get current width
        curHeight = root.winfo_height()  # get current height
        scnWidth, scnHeight = root.maxsize()  # get screen width and height
        tmpcnf = '+%d+%d' % ((scnWidth - curWidth) / 2, (scnHeight - curHeight) / 2)
        root.geometry(tmpcnf)
        root.mainloop()

    def create_menu(self,root):
        #创建菜单栏
        menu = Menu(root)
        #创建二级菜单
        file_menu = Menu(menu,tearoff=0)
        file_menu.add_command(label="设置路径",command=self.open_dir)
        file_menu.add_separator()
        scan_menu = Menu(menu)
        file_menu.add_command(label="扫描")
        about_menu = Menu(menu,tearoff=0)
        about_menu.add_command(label="version:1.0")
        #在菜单栏中添加以下一级菜单
        menu.add_cascade(label="文件",menu=file_menu)
        menu.add_cascade(label="关于",menu=about_menu)
        root['menu'] = menu

    def create_content(self, root):
        lf = ttk.LabelFrame(root, text="文件搜索")
        lf.pack(fill=X, padx=15, pady=8)
        top_frame = Frame(lf)
        top_frame.pack(fill=X,expand=YES,side=TOP,padx=15,pady=8)
        self.search_key = StringVar()
        ttk.Entry(top_frame, textvariable=self.search_key,width=50).pack(fill=X,expand=YES,side=LEFT)
        ttk.Button(top_frame,text="搜索",command=self.search_file).pack(padx=15,fill=X,expand=YES)
        bottom_frame = Frame(lf)
        bottom_frame.pack(fill=BOTH,expand=YES,side=TOP,padx=15,pady=8)

        band = Frame(bottom_frame)
        band.pack(fill=BOTH,expand=YES,side=TOP)

        self.list_val = StringVar()
        listbox = Listbox(band,listvariable=self.list_val,height=18)
        listbox.pack(side=LEFT,fill=X,expand=YES)

        vertical_bar = ttk.Scrollbar(band,orient=VERTICAL,command=listbox.yview)
        vertical_bar.pack(side=RIGHT,fill=Y)
        listbox['yscrollcommand'] = vertical_bar.set

        horizontal_bar = ttk.Scrollbar(bottom_frame,orient=HORIZONTAL,command=listbox.xview)
        horizontal_bar.pack(side=BOTTOM,fill=X)
        listbox['xscrollcommand'] = horizontal_bar.set

        #给list动态设置数据，set方法传入一个元组，注意此处是设置，不是插入数据，此方法调用后，list之前的数据会被清除
        self.list_val.set(('jioj',124,"fjoweghpw",1,2,3,4,5,6))

    def search_file(self):
        pass

    def open_dir(self):
        d = dir.Directory()
        self.path = d.show(initialdir=self.path)
'''
具体调用：
if __name__ == "__main__":
    AppUI()
'''
#****窗口类2********************************************************************
import tkinter as tk 
import tkinter.scrolledtext as tst

#v = tk.StringVar()
class Application(tk.Frame):
    def __init__(self, master=None):  
        tk.Frame.__init__(self, master)
        self.grid()
        self.createWidgets()     
        self.createMenu()         
        root['menu'] = self.menubar 
        root.bind('<Button-3>', self.f_popup)
        #root.bind('<Button-1>', self.callback)

         # 状态栏
        self.status = tk.StringVar()
        self.status.set('Ln: 1 Col: 1')
        self.lblStatus = tk.Label(self, textvariable=self.status, anchor='c')
        self.lblStatus.grid(row=7, column=0, columnspan=20, sticky = tk.S + tk.E)
        root.bind('<Key>', self.loc)
        root.bind('<Button-1>', self.loc)
        
    def createWidgets(self):             
        self.textEdit = tst.ScrolledText(self, width=80, height=25)
        self.textEdit.grid(row=0, column=0, rowspan=6)
        
        '''
        self.lblState = tk.Label(self, text="状态栏:")
        self.lblState.grid(row=7, column=0, columnspan=10)
        self.entryState = tk.Entry(self, textvariable=v)
        self.lblState.grid(row=7, column=11, columnspan=20)
        '''
        
    def createMenu(self):              
        self.menubar = tk.Menu(root)   
        #创建子菜单
        self.menufile  = tk.Menu(self.menubar)     
        self.menuedit = tk.Menu(self.menubar, tearoff=0) 
        self.menuhelp = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label='File', menu=self.menufile)
        self.menubar.add_cascade(label="Edit", menu=self.menuedit)
        self.menubar.add_cascade(label="Help", menu=self.menuhelp)
        
        #添加菜单项
        self.menufile.add_command(label='New', command=self.f_new)
        self.menufile.bind("<Control-N>", self.f_new)
        self.menufile.add_command(label='Open', command=self.f_open)
        self.menufile.bind("<Control-O>", self.f_open)
        self.menufile.add_command(label='Save', accelerator='^A',command=self.f_save)
        self.menufile.bind("<Control-S>", self.f_save)
        #分隔
        self.menufile.add_separator()  
        self.menufile.add_command(label='Exit', command=root.destroy) 
        self.menuedit.add_command(label="Cut", command=self.f_cut, accelerator='Ctrl+T')
        self.menuedit.bind("<Control-T>", self.f_cut)
        self.menuedit.add_command(label="Copy", command=self.f_copy, accelerator='Ctrl+C')
        self.menuedit.bind("<Control-T>", self.f_copy)
        self.menuedit.add_command(label="Paste", command=self.f_paste, accelerator='Ctrl+V')
        self.menuedit.bind("<Control-T>", self.f_paste)
        self.menuedit.add_command(label="Delete", command=self.f_delete, accelerator='Ctrl+D')
        self.menuedit.bind("<Control-T>", self.f_delete)
        self.menuhelp.add_command(label="About", command=self.f_about)
    def f_new(self):
        root.title('untitle')       #新建文件title
        self.textEdit.delete(1.0, tk.END)  
    def f_open(self):           
        self.textEdit.delete(1.0, tk.END) 
        fname = tk.filedialog.askopenfilename(filetypes=[('Python源文件','.py')])
        with open(fname, 'r', encoding= 'utf-8') as f1: 
            str1 = f1.read()                
        self.textEdit.insert(0.0, str1)         
    def f_save(self):          
        str1 = self.textEdit.get(1.0, tk.END)   
        fname = tk.filedialog.asksaveasfilename(filetypes=[('Python源文件','.py')])
        with open(fname, 'w', encoding= 'utf-8') as f1:  #打开文件
            f1.write(str1)      
    def f_about(self):           
        tk.messagebox.showinfo('关于', '版本V 1.0.1')
    def f_cut(self):
        try:
            str1=self.textEdit.get(tk.SEL_FIRST, tk.SEL_LAST)
            self.textEdit.clipboard_clear()        
            self.textEdit.clipboard_append(str1)    
            self.textEdit.delete(tk.SEL_FIRST, tk.SEL_LAST)
        except: pass
    def f_copy(self):           
        try:
            str1=self.textEdit.get(tk.SEL_FIRST, tk.SEL_LAST)
            self.textEdit.clipboard_clear()        
            self.textEdit.clipboard_append(str1)   
        except: pass
    def f_paste(self):          
        str1 = self.textEdit.selection_get(selection='CLIPBOARD')
        try:           
            self.textEdit.replace(tk.SEL_FIRST, tk.SEL_LAST, str1)
        except:
            self.textEdit.insert(tk.INSERT, str1)
    def f_delete(self):          
        try:
            str1=self.textEdit.get(tk.SEL_FIRST, tk.SEL_LAST) 
            self.textEdit.delete(tk.SEL_FIRST, tk.SEL_LAST)
        except: pass
    #定位光标位置，缺点读字符串的位置。。，不知道换行在哪里，仍需加强条件调试。
    def loc(self,event=None):
        s = str(self.textEdit.index('insert')).split('.')
        statusXY = 'Ln: ' + str(s[0]) + 'Col: ' + str(s[1])
        self.status.set(statusXY)

    def f_popup(self, event):   
        self.menuedit.post(event.x_root, event.y_root)
'''
#具体调用：
root = tk.Tk()               
root.title('文本编辑器')
app = Application(master=root)   
app.mainloop()
'''

#****返回文件名对话框======================================
def openDialog():
  import tkinter as tk
  from tkinter import filedialog #此句有个Tk窗口，处于未响应状态，故隐藏之
  root = tk.Tk()
  root.withdraw() #隐藏之
  file_path = filedialog.askopenfilename() #注意是name结尾不是names
  return file_path

#=============================非界面框架========================================
#*******************************************************************************

#*******************************************************************
#工作簿导出为无公式工作簿；另存为："无公式"+原文件 的新文件名
def excel无公式导出(bookPath=''):
	oExcel=vbaInit()
	if bookPath=='': filePath=oExcel.Application.FileDialog(1)
	obj=openEx(bookPath)
	for sht in obj.Sheets:
		sht.UsedRange.Value = sht.UsedRange.Value
	obj.Close(True,obj.Path+"\\无公式"+obj.Name)
	tkinter.messagebox.showinfo('','导出成功!')
	
#*********************************************************************
#解压单个zip和rar文件
def 解压缩文件(file_name=""): #解压至文件名+后缀_的目录里;
	import rarfile; import zipfile
	if file_name=="": file_name=openDialog()  
	path=file_name+"_"
	if not os.path.isdir(path):
		os.mkdir(path)
	if file_name.endswith(".rar"):
		try:
			rar = rarfile.RarFile(file_name)
			rar.extractall(path)
			rar.close()
			return 1
		except Exception as e: print('解压缩文件[{}]失败'.format(file_name), str(e))
	elif file_name.endswith(".zip"):
		try:
			zip_file = zipfile.ZipFile(file_name)
			for names in zip_file.namelist():    #??不用循环怎么处理
				zip_file.extract(names, path)
			zip_file.close()
			return 1
		except Exception as e: print('解压失败'.format(file_name), str(e))
	return 0

#****解压文件夹内所有文件*************************************************
def 解压缩文件夹(folder="",delete=1): #delete: 1删除压缩包，0不删除
	if folder=="":	folder=askdirectory()
	files=list_all_files(folder)
	#tkinter.messagebox.showinfo("",files)
	for file in files:
		sv=解压缩文件(file)
		if sv and delete: os.remove(file)  #删除源文件，sv!=0 不会误删其他非压缩文件
	#tkinter.messagebox.showinfo('end!')

#列出文件夹及子文件夹下所有文件********************************************
def list_all_files(rootdir=""):   
	_files = []
	if rootdir=="":  rootdir=askdirectory()
	list = os.listdir(rootdir) 
	for i in range(0,len(list)):
		path = os.path.join(rootdir,list[i])
		if os.path.isdir(path):  _files.extend(list_all_files(path))
		if os.path.isfile(path) and not '~$' in path:  _files.append(path) #过滤了临时文件
	return _files

#****读取js文件中的函数，但是对微软的Jscript无效
def js_eval(js,func,*para):
  import execjs
  ff= execjs.compile(open(js).read()).call(func,*para) #
  


#****获取文件的md5值**********************************************
def get_md5(filename): 
  m = hashlib.md5() 
  mfile = open(filename, "rb") 
  m.update(mfile.read()) 
  mfile.close() 
  md5_value = m.hexdigest() 
  return md5_value 


#***********************+******************************************
def 删除重复文件_cmp(path='',num=1): #num目录个数
	all_files=[]
	if path=='': 
		dirA=askdirectory()+"\\" ##获取目录名，删除相同文件是不删这个目录里文件的
		list1 = list_all_files(dirA)
		all_files.extend(list1)
	if num>1:
		for i in range(num-1):
			list1 = list_all_files(path)
			all_files.extend(list1)
	print(dirA)
	#用双重for循环，令每个文件都与全部文件进行一次对比
	
	for x in all_files:
		for y in all_files:
			if x != y and os.path.exists(x) and os.path.exists(y):
				if filecmp.cmp(x, y):
					if dirA not in y and num>1: os.remove(y) #此句确保不删dirA目录文件
					elif num==1: os.remove(y)

#****去重**********************************************************
def 删除重复文件_md5(path=""):
	if path=="": path=askdirectory()
	delstr=[]
	解压缩文件夹(path)
	md5List =[] 
	urlList =list_all_files(path) 
	for a in urlList: 
		md5 =get_md5(a) 
		if (md5 in md5List): 
			os.remove(a)
			delstr.append(a) 
		else: 
			md5List.append(md5)
	删除空文件夹(path)  #删除所有空文件夹
	日志(delstr)
	
#****删除一个文件夹里的所有空文件夹********************************
def 删除空文件夹(path=''):
	if path=="": path=askdirectory()
	dir_list=[]
	for root,dirs,files in os.walk(path): #遍历文件夹
	    dir_list.append(root)
	# 先生成文件夹的列表，重点是下边
	for root in dir_list[::-1]:
	    if not os.listdir(root): os.rmdir(root)
	
#******************************************************************
#win10桌面动态路径
def 桌面路径():	
	key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
	return winreg.QueryValueEx(key,"Desktop")[0]

#*******************************************************************
#读取一列转换为list
def excel_col_to_list(file,col=[1]): #注意第二列是 [1]，第一列是[0];names=None不要列名
    df = pd.read_excel(file,usecols=col,names=None)  
    df_li = df.values.tolist()
    result = []
    for s_li in df_li:
        result.append(s_li[0]) #提取
    return result
   
#******************************************************************
#
def 日志(sss='',mode='a',f="日志.txt"):
	f1 = open(f,mode)
	if type(sss) == list: #list类型加入换行，看起来清晰
		for i in sss:
			f1.write(i+"\n")
	else: f1.write(sss+"\n")
	f1.close()


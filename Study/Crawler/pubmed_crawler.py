import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import os
 
def geturl(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        print('shibai')
 
def parserinfo(infolist, html, keyword):
    infonums = re.findall(r'\"resultcount\" value=\"[\d\.]*', html) #(r'Items: 1 to [\d\.]* of [\d\.]*', html)
    for infonum in infonums:
        info = infonum.split("value=\"")[1]
        infolist.append([keyword,info])
 
def printinfolist(info):
    for i, item in enumerate(info):
        print(i+1, item[0], item[1])
 
def read_data(inputfile, infolist):
    wb = load_workbook(inputfile)
    sheet = wb["Sheet1"] #表格中的sheet1
    for i in sheet["A"]: #A列
        if i is not sheet["A1"] and i.value is not None:
            infolist.append(i.value)
 
def saverinfo(outputfile, info):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "New shit"
    sheet['A1'] = "编号"
    sheet['B1'] = "名称"
    sheet['C1'] = "数目"
    for i, item in enumerate(info):
        sheet['A%d' % (i+2)].value = str(i+1)
        sheet['B%d' % (i+2)].value = item[0]
        sheet['C%d' % (i+2)].value = item[1]
    wb.save(outputfile)
 
def main():
    starturl = "https://www.ncbi.nlm.nih.gov/pubmed?term=" #网址接口
    keyword = []
    info = []
    path = os.getcwd()
    inputfile = path + "/miRNA.xlsx" #输入表格名字
    read_data(inputfile, keyword)
    outputfile = path + "/test1.xlsx"
    count = len(keyword)
    for i, key in enumerate(keyword):
        url = starturl + key +str("%5BTitle%2FAbstract%5D")
        html = geturl(url)
        parserinfo(info, html, key)
        print('%d' % i + '%' + str(count))
    saverinfo(outputfile, info)
 
main()
